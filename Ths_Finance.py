import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
import json
import time
import threading

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class TongHuaShun(object):
    def __init__(self, file_name):
        self.xlsxname = file_name

    def request(self):
        self.url = 'http://www.10jqka.com.cn/'
        self.data = requests.get(self.url, headers=headers)
        self.data.encoding = "gbk"
        self.soup = BeautifulSoup(self.data.text, "lxml")

        #calendar
        calendar_time = time.strftime("%Y%m", time.localtime())  # year-month-day-hour-minute
        #self.url_calendar = 'http://stock.10jqka.com.cn/fincalendar.shtml#{}'.format(datatime)
        self.url_calendar = 'http://comment.10jqka.com.cn/tzrl/getTzrlData.php?callback=callback_dt&type=data&date={}'.format(calendar_time)
        self.data_calendar = requests.get(self.url_calendar, headers=headers)

    def Style(self):
        self.m_font = Font(
            size=12,
            bold=True,
        )

        self.head_font = Font(
            size=14,
            bold=True,
        )

    def getNew(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet('Ths')
        datalist = self.soup.find_all(class_="item_txt")
        t_row = 1
        t_col = 1

        sheet.cell(row=t_row + 0, column=t_col + 0, value="同花顺财经")
        sheet.cell(row=t_row + 1, column=t_col + 0, value="新闻标题")
        sheet.cell(row=t_row + 1, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row + 1, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row + 1, column=t_col + 3, value="新闻时间")
        t_row = t_row + 2

        for news in datalist:
            newlist2 = news.select('p a')
            for m_new in newlist2:
                m_url = m_new['href']
                m_title = m_new['title']
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("THS Save Error = 1")

    def getInvestment(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Ths')
        t_row = sheet.max_row

        datalist = self.soup.find_all(class_="content newhe") #投资机会上半部分 产经新闻 研报精选
        t_col = 1
        index = 0
        for newlist in datalist:
            #里面有个重复 加个判断去掉
            news = newlist.select('li a')
            for m_new in news:
                if index != 1:
                    m_url = m_new['href']
                    m_title = m_new['title']
                    sheet.cell(row=t_row, column=t_col, value=m_title)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                    t_row = t_row + 1
                index = index + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("THS Save Error = 2")

    def getInvestment2(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Ths')
        t_row = sheet.max_row
        t_col = 1
        #投资机会后半部分获取
        datalist = self.soup.find_all(class_="last")

        i = 0
        for newlist in datalist:
            news = newlist.select('li a')
            for m_new in news:
                #这个分类筛选的结果重复的太多  但是数量是固定的
                if i >= 5:
                    if i < 11:
                        m_title = m_new.get_text()
                        m_url = m_new['href']
                        sheet.cell(row=t_row, column=t_col, value=m_title)
                        sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                        t_row = t_row + 1
                i = i + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("THS Save Error = 3")

    def get_Newspaper(self):
        url_Newspaper = 'http://stock.10jqka.com.cn/bktt_list/'
        data_paper = requests.get(url_Newspaper, headers=headers)
        soup_paper = BeautifulSoup(data_paper.text, "lxml")

        datalist = soup_paper.select('body > div.content-1200 > div.module-l.fl > div.list-con > ul > li:nth-child(1) > span > a')
        datalist2 = soup_paper.select('body > div.content-1200 > div.module-l.fl > div.list-con > ul > li:nth-child(1) > a')
        m_url = datalist[0]['href']
        m_title = datalist[0]['title']
        m_content = datalist2[0].get_text() + "del"
        m_content = m_content.replace("...del", "")

        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Ths')
        t_row = sheet.max_row + 1
        t_col = 1

        sheet.cell(row=t_row + 1, column=t_col, value="报刊头条") #下一个函数省一次xlwt操作 (投资日历)
        t_row = t_row + 2
        sheet.cell(row=t_row, column=t_col, value=m_title)
        sheet.cell(row=t_row, column=t_col + 1, value=m_url)
        sheet.cell(row=t_row, column=t_col + 2, value=m_content)

        # 为下一个函数省一次xlwt操作 (投资日历)
        sheet.cell(row=t_row + 2, column=t_col, value="投资日历")
        sheet.cell(row=t_row + 3 ,column=t_col, value="会议事件")
        sheet.cell(row=t_row + 3, column=t_col + 1, value="会议地点")
        sheet.cell(row=t_row + 3, column=t_col + 2, value="会议时间")
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("THS Save Error = 4")

    threadLock = threading.Lock()
    def dealjson(self, json): #时间 事件 地点 json问题相关板块顺序是乱序 就不加了
        self.threadLock.acquire()
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Ths')
        t_row = sheet.max_row + 1
        t_col = 1

        for json_event in json['events']:
            m_event = json_event[0]
            m_place = json_event[2]
            sheet.cell(row=t_row, column=t_col, value=m_event)
            sheet.cell(row=t_row, column=t_col + 1, value=m_place)
            m_date = json['date']
            m_week = json['week']
            sheet.cell(row=t_row, column=t_col + 2, value=m_date + "-" + m_week)
            t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
            self.threadLock.release()
        except Exception:
            self.threadLock.release()
            print("THS Save Error = 5")


    def get_Calendar(self): #投资日历
        data = self.data_calendar.text + "del"
        data = data.replace("callback_dt(", "")
        data = data.replace(");del", "")
        json_str = json.loads(data)
        t = 0
        #pool = ThreadPoolExecutor(max_workers=2)
        for m_json in json_str['data']:
            #future1 = pool.submit(self.dealjson, m_json)
            t1 = threading.Thread(target=self.dealjson, args=(m_json, ))
            t1.start()
            t1.join()

    def main(self, file_name):
        Ths = TongHuaShun(file_name)
        Ths.request()
        Ths.Style()
        Ths.getNew()
        Ths.getInvestment()
        Ths.getInvestment2()
        Ths.get_Newspaper()
        Ths.get_Calendar()
