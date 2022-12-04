import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import json
import threading
import time
import re
import os
import wget
from src.Platform import pt
#处理个股版本，包含交易明细、买卖盘、券商机构等。
#上层目录Code.txt 添加代码
#失效不再更新，自行抓包获取。


headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class SelfStock(object):
    threadLock = threading.Lock()
    def __init__(self):
        self.stop = True
        #self.xlsxname = file_name

    def Deal_Xq_quote(self, data, name):
        self.threadLock.acquire()
        file_name = self.get_filename(name)
        try:
            wb = load_workbook(file_name)
            try:
                sheet = wb.get_sheet_by_name(name)
                ws = wb[name]
                wb.remove(ws)
                sheet = wb.create_sheet(name, 0)
            except:
                sheet = wb.create_sheet(name)
        except:
            wb = Workbook()
            ws = wb['Sheet']
            wb.remove(ws)
            sheet = wb.create_sheet(name)

        t_row = 1
        t_col = 1

        data_json = data['data']
        data_items = data_json['items']
        data_mk = data_items[0]['market']
        data_quote = data_items[0]['quote']
        #print(data_quote)

        sheet.cell(row=t_row, column=t_col, value="股票代码")
        sheet.cell(row=t_row, column=t_col + 1, value="股票名称")
        sheet.cell(row=t_row, column=t_col + 2, value="交易状态")
        sheet.cell(row=t_row, column=t_col + 3, value="更新时间")
        t_row = t_row + 1
        m_status = data_mk['status']
        stock_code = data_quote['symbol']
        stock_name = data_quote['name']
        m_time = data_quote['timestamp']
        timeStamp = float(m_time / 1000) #13位时间戳
        timeArray = time.localtime(timeStamp)
        otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)

        sheet.cell(row=t_row, column=t_col, value=stock_code)
        sheet.cell(row=t_row, column=t_col + 1, value=stock_name)
        sheet.cell(row=t_row, column=t_col + 2, value=m_status)
        sheet.cell(row=t_row, column=t_col + 3, value=otherStyleTime)
        t_row = t_row + 2

        sheet.cell(row=t_row, column=t_col + 0, value="当前价格")
        sheet.cell(row=t_row, column=t_col + 1, value="涨跌价格")
        sheet.cell(row=t_row, column=t_col + 2, value="涨跌幅度")
        sheet.cell(row=t_row, column=t_col + 3, value="开盘价格")
        sheet.cell(row=t_row, column=t_col + 4, value="当前新高")
        sheet.cell(row=t_row, column=t_col + 5, value="当前新低")
        sheet.cell(row=t_row, column=t_col + 6, value="昨日收盘")
        sheet.cell(row=t_row, column=t_col + 7, value="平均价格")
        sheet.cell(row=t_row, column=t_col + 8, value="涨停价格")
        sheet.cell(row=t_row, column=t_col + 9, value="跌停价格")
        sheet.cell(row=t_row, column=t_col + 10, value="52周最高")
        sheet.cell(row=t_row, column=t_col + 11, value="52周最低")
        t_row = t_row + 1

        m_current = data_quote['current']
        m_percent = data_quote['percent'] #涨跌幅度
        m_chg = data_quote['chg'] #涨跌价格
        m_open = data_quote['open'] #开盘价
        m_yesclose = data_quote['last_close'] #昨收
        m_high = data_quote['high']
        m_low = data_quote['low']
        m_avg_price = data_quote['avg_price']
        m_limit_up = data_quote['limit_up'] #涨停
        m_limit_down = data_quote['limit_down'] #跌停
        m_high52w = data_quote['high52w'] #52周最高
        m_low52w = data_quote['low52w'] #52周最低
        sheet.cell(row=t_row, column=t_col + 0, value=m_current)
        sheet.cell(row=t_row, column=t_col + 1, value=m_chg)
        sheet.cell(row=t_row, column=t_col + 2, value=str(m_percent) + "%")
        sheet.cell(row=t_row, column=t_col + 3, value=m_open)
        sheet.cell(row=t_row, column=t_col + 4, value=m_high)
        sheet.cell(row=t_row, column=t_col + 5, value=m_low)
        sheet.cell(row=t_row, column=t_col + 6, value=m_yesclose)
        sheet.cell(row=t_row, column=t_col + 7, value=m_avg_price)
        sheet.cell(row=t_row, column=t_col + 8, value=m_limit_up)
        sheet.cell(row=t_row, column=t_col + 9, value=m_limit_down)
        sheet.cell(row=t_row, column=t_col + 10, value=m_high52w)
        sheet.cell(row=t_row, column=t_col + 11, value=m_low52w)
        t_row = t_row + 2


        sheet.cell(row=t_row, column=t_col + 0, value="成交额(/万)")
        sheet.cell(row=t_row, column=t_col + 1, value="成交量(/万手)")
        sheet.cell(row=t_row, column=t_col + 2, value="换手率")
        sheet.cell(row=t_row, column=t_col + 3, value="量比")
        sheet.cell(row=t_row, column=t_col + 4, value="振幅")
        sheet.cell(row=t_row, column=t_col + 5, value="市盈率TTM")
        sheet.cell(row=t_row, column=t_col + 6, value="市盈率(动)")
        sheet.cell(row=t_row, column=t_col + 7, value="市盈率(静)")
        sheet.cell(row=t_row, column=t_col + 8, value="市净率")
        sheet.cell(row=t_row, column=t_col + 9, value="总股本(/万)")
        sheet.cell(row=t_row, column=t_col + 10, value="流通股(/万)")
        sheet.cell(row=t_row, column=t_col + 11, value="总市值(/亿)")
        sheet.cell(row=t_row, column=t_col + 12, value="流通市值(/亿)")
        t_row = t_row + 1

        m_amount = data_quote['amount'] #成交额
        if m_amount == None: #防止停牌、退市等情况股价不再更新
            self.stop = False
            self.threadLock.release()
            return
        m_turnover_rate = data_quote['turnover_rate'] #换手：0.74%
        m_volume = data_quote['volume'] #成交量
        m_amplitude = data_quote['amplitude'] #振幅
        m_total_shares = data_quote['total_shares'] #总股本
        m_float_shares = data_quote['float_shares'] #流通股
        m_volume_ratio = data_quote['volume_ratio'] #量比
        m_pe_ttm = data_quote['pe_ttm'] #市盈率
        m_pe_forecast = data_quote['pe_forecast'] #动态市盈率
        m_pe_lyr = data_quote['pe_lyr'] #静态市盈率
        m_pb = data_quote['pb'] #市净率
        m_profit = data_quote['profit'] #年报利润
        m_profit_four = data_quote['profit_four'] #也是利润 不清楚
        m_profit_forecast = data_quote['profit_forecast']
        m_market_capital = data_quote['market_capital'] #总市值
        m_float_market_capital = data_quote['float_market_capital'] #流通市值
        sheet.cell(row=t_row, column=t_col + 0, value=round(m_amount / 10000, 2))
        sheet.cell(row=t_row, column=t_col + 1, value=round(m_volume / 10000, 2))
        sheet.cell(row=t_row, column=t_col + 2, value=str(m_turnover_rate) + "%")
        sheet.cell(row=t_row, column=t_col + 3, value=m_volume_ratio)
        sheet.cell(row=t_row, column=t_col + 4, value=str(m_amplitude) + "%")
        sheet.cell(row=t_row, column=t_col + 5, value=m_pe_ttm)
        sheet.cell(row=t_row, column=t_col + 6, value=m_pe_forecast)
        sheet.cell(row=t_row, column=t_col + 7, value=m_pe_lyr)
        sheet.cell(row=t_row, column=t_col + 8, value=m_pb)
        sheet.cell(row=t_row, column=t_col + 9, value=round(m_total_shares / 10000, 2))
        sheet.cell(row=t_row, column=t_col + 10, value=round(m_float_shares / 10000, 2))
        sheet.cell(row=t_row, column=t_col + 11, value=round(m_market_capital / 100000000, 2))
        sheet.cell(row=t_row, column=t_col + 12, value=round(m_float_market_capital / 100000000, 2))

        self.threadLock.release()

        try:
            wb.save(file_name)
        except Exception:
            print("Self_Stock Save Error = Xq_quote")

    def Deal_Xq_distribution(self, data, name):
        self.threadLock.acquire()
        file_name = self.get_filename(name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name(name)

        t_row = sheet.max_row + 3
        t_col = 1

        m_text = data['data']['analysis'][0] #今日主力净流入XX亿
        sheet.cell(row=t_row, column=t_col, value=m_text)
        t_row = t_row + 1

        sheet.cell(row=t_row, column=t_col, value="资金成交分布(/万)")
        sheet.cell(row=t_row, column=6, value="净流入(/万)")
        t_row = t_row + 1

        m_data = data['data']['distribution']
        m_sell = m_data['sell'] #
        m_buy = m_data['buy']

        sheet.cell(row=t_row, column=t_col, value="特大单卖出")
        sheet.cell(row=t_row + 1, column=t_col, value="大单卖出")
        sheet.cell(row=t_row + 2, column=t_col, value="中单卖出")
        sheet.cell(row=t_row + 3, column=t_col, value="小单卖出")
        sheet.cell(row=t_row + 4, column=t_col, value="合计")
        se_xlarge = m_sell['xlarge']
        se_large = m_sell['large']
        se_medium = m_sell['medium']
        se_small = m_sell['small']
        sum_sell = se_large + se_large + se_medium + se_small

        by_xlarge = m_buy['xlarge']
        by_large = m_buy['large']
        by_medium = m_buy['medium']
        by_small = m_buy['small']
        sum_buy = by_xlarge + by_large + by_medium + by_small
        t_col = t_col + 1
        sheet.cell(row=t_row, column=t_col, value=round(se_xlarge / 10000, 2))
        sheet.cell(row=t_row + 1, column=t_col, value=round(se_large / 10000, 2))
        sheet.cell(row=t_row + 2, column=t_col, value=round(se_medium / 10000, 2))
        sheet.cell(row=t_row + 3, column=t_col, value=round(se_small / 10000, 2))
        sheet.cell(row=t_row + 4, column=t_col, value=round(sum_sell / 10000, 2))
        t_col = t_col + 2

        sheet.cell(row=t_row, column=t_col, value=round(by_xlarge / 10000, 2))
        sheet.cell(row=t_row + 1, column=t_col, value=round(by_large / 10000, 2))
        sheet.cell(row=t_row + 2, column=t_col, value=round(by_medium / 10000, 2))
        sheet.cell(row=t_row + 3, column=t_col, value=round(by_small / 10000, 2))
        sheet.cell(row=t_row + 4, column=t_col, value=round(sum_buy / 10000, 2))
        t_col = t_col + 1

        sheet.cell(row=t_row, column=t_col, value="特大单买入")
        sheet.cell(row=t_row + 1, column=t_col, value="大单买入")
        sheet.cell(row=t_row + 2, column=t_col, value="中单买入")
        sheet.cell(row=t_row + 3, column=t_col, value="小单买入 ")
        sheet.cell(row=t_row + 4, column=t_col, value="净流入 ")

        m_xlarge = round((by_xlarge / 10000) - (se_xlarge / 10000), 2)
        m_large = round((by_large / 10000) - (se_large / 10000), 2)
        m_medium = round((by_medium / 10000) - (se_medium / 10000), 2)
        m_small = round((by_small / 10000) - (se_small / 10000), 2)
        sheet.cell(row=t_row, column=6, value=m_xlarge)
        sheet.cell(row=t_row + 1, column=6, value=m_large)
        sheet.cell(row=t_row + 2, column=6, value=m_medium)
        sheet.cell(row=t_row + 3, column=6, value=m_small)
        sheet.cell(row=t_row + 4, column=6, value=round(sum_buy - sum_sell, 2) / 10000)

        self.threadLock.release()

        try:
            wb.save(file_name)
        except Exception:
            print("Self_Stock Save Error = distribution")


    def Deal_Xq_query(self, data, name): #主力历史是个持续更新的东西 保存到一个新的文件中 分为第一次使用或者长期更新
        self.threadLock.acquire()
        file_name = self.get_filename(name)
        wb = load_workbook(file_name)
        new_sheet = False
        try:
            sheet = wb.get_sheet_by_name("资金流向历史")
        except:
            sheet = wb.create_sheet("资金流向历史")
            new_sheet = True
            sheet.cell(row=1, column=1, value="日期")
            sheet.cell(row=1, column=2, value="收盘价")
            sheet.cell(row=1, column=3, value="涨跌幅")
            sheet.cell(row=1, column=4, value="主力资金净流入")
            sheet.cell(row=1, column=5, value="特大单净流入")
            sheet.cell(row=1, column=6, value="大单净流入")
            sheet.cell(row=1, column=7, value="中单净流入")
            sheet.cell(row=1, column=8, value="小单净流入")

        m_data = data['data']['items']
        new_max_line = 20 + 1  # 标题一行 内容20行
        t_row = sheet.max_row + 1
        n = 0
        t_col = 1
        for m_json in m_data:
            m_small = m_json['small']
            m_large = m_json['large']
            m_xlarge = m_json['xlarge']
            m_medium = m_json['medium']
            m_close = m_json['close']  # 收盘价
            m_percent = m_json['percent']  # 涨跌幅
            m_amount = m_json['amount']  # 主力
            m_time = m_json['timestamp']
            timeStamp = float(m_time / 1000)  # 13位时间戳
            timeArray = time.localtime(timeStamp)
            otherStyleTime = time.strftime("%Y-%m-%d", timeArray)  # 年月日

            if new_sheet == False:
                cell = sheet.cell(t_row - 1, 1).value
                if cell == otherStyleTime:
                    t_row = t_row - 1
                    sheet.cell(row=t_row, column=t_col, value=otherStyleTime)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_close)
                    sheet.cell(row=t_row, column=t_col + 2, value=str(m_percent) + "%")
                    sheet.cell(row=t_row, column=t_col + 3, value=round(m_amount / 10000, 2))
                    sheet.cell(row=t_row, column=t_col + 4, value=round(m_xlarge / 10000, 2))
                    sheet.cell(row=t_row, column=t_col + 5, value=round(m_large / 10000, 2))
                    sheet.cell(row=t_row, column=t_col + 6, value=round(m_medium / 10000, 2))
                    sheet.cell(row=t_row, column=t_col + 7, value=round(m_small / 10000, 2))
                    break
                elif n < 2:
                    sheet.cell(row=t_row, column=t_col, value=otherStyleTime)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_close)
                    sheet.cell(row=t_row, column=t_col + 2, value=str(m_percent) + "%")
                    sheet.cell(row=t_row, column=t_col + 3, value=round(m_amount / 10000, 2))
                    sheet.cell(row=t_row, column=t_col + 4, value=round(m_xlarge / 10000, 2))
                    sheet.cell(row=t_row, column=t_col + 5, value=round(m_large / 10000, 2))
                    sheet.cell(row=t_row, column=t_col + 6, value=round(m_medium / 10000, 2))
                    sheet.cell(row=t_row, column=t_col + 7, value=round(m_small / 10000, 2))
                    t_row = t_row - 1
                    n = n + 1
                else:
                    break
            elif new_sheet == True:  # 新文件/新表 写入全部日期
                if new_max_line > 1:
                    sheet.cell(row=new_max_line, column=t_col, value=otherStyleTime)
                    sheet.cell(row=new_max_line, column=t_col + 1, value=m_close)
                    sheet.cell(row=new_max_line, column=t_col + 2, value=str(m_percent) + "%")
                    sheet.cell(row=new_max_line, column=t_col + 3, value=round(m_amount / 10000, 2))
                    sheet.cell(row=new_max_line, column=t_col + 4, value=round(m_xlarge / 10000, 2))
                    sheet.cell(row=new_max_line, column=t_col + 5, value=round(m_large / 10000, 2))
                    sheet.cell(row=new_max_line, column=t_col + 6, value=round(m_medium / 10000, 2))
                    sheet.cell(row=new_max_line, column=t_col + 7, value=round(m_small / 10000, 2))
                    new_max_line = new_max_line - 1

        self.threadLock.release()

        try:
            wb.save(file_name)
        except:
            print("Self_Stock Save Error = query")

        """
        flag = True #已经存在为true
        new_sheet = True #新sheet

        if if_os == True:
            desktop_path = os.path.join(os.path.expanduser('~'), "Desktop")  # 获取桌面路径
            path = desktop_path + "\\Finance\\History\\闻讯_主力资金历史.xlsx"
            isExists = os.path.exists(path)
            if not isExists:
                flag = False
                new_sheet = True
                if self.first == 0:
                    self.wb = Workbook()
                    ws = self.wb['Sheet']
                    self.wb.remove(ws)
                    self.sheet = self.wb.create_sheet(name)
                    self.first += 1
                else:
                    self.wb = load_workbook(desktop_path + "\\Finance\\Main_History.xlsx")
                    try:
                        self.sheet = self.wb.get_sheet_by_name(name)
                        new_sheet = False
                    except:
                        self.sheet = self.wb.create_sheet(name)
                        new_sheet = True
            else:
                self.wb = load_workbook(path)
                try:
                    self.sheet = self.wb.get_sheet_by_name(name)
                    new_sheet = False
                except:
                    self.sheet = self.wb.create_sheet(name)
                    new_sheet = True
        elif if_os == False:
            d_path = "./Finance/History/"
            paths = d_path + "闻讯_主力资金历史.xlsx"
            isExists = os.path.exists(paths)
            if not isExists:
                flag = False #不存在
                new_sheet = True
                if self.first == 0:
                    self.wb = Workbook()
                    ws = self.wb['Sheet']
                    self.wb.remove(ws)
                    self.sheet = self.wb.create_sheet(name)
                    self.first += 1
                else:
                    self.wb = load_workbook("./Main_History.xlsx")
                    try:
                        self.sheet = self.wb.get_sheet_by_name(name)
                        new_sheet = False
                    except:
                        self.sheet = self.wb.create_sheet(name)
                        new_sheet = True
            else:
               self.wb = load_workbook(paths)
               try:
                    self.sheet = self.wb.get_sheet_by_name(name)
                    new_sheet = False
               except:
                    self.sheet = self.wb.create_sheet(name)
                    new_sheet = True

        if new_sheet == True:
            self.sheet.cell(row=1, column=1, value="日期")
            self.sheet.cell(row=1, column=2, value="收盘价")
            self.sheet.cell(row=1, column=3, value="涨跌幅")
            self.sheet.cell(row=1, column=4, value="主力资金净流入")
            self.sheet.cell(row=1, column=5, value="特大单净流入")
            self.sheet.cell(row=1, column=6, value="大单净流入")
            self.sheet.cell(row=1, column=7, value="中单净流入")
            self.sheet.cell(row=1, column=8, value="小单净流入")

        max_row = 20 + 1 #标题一行 内容20行
        n = 0
        t_row = self.sheet.max_row + 1
        t_col = 1
        for m_json in m_data:
            m_small = m_json['small']
            m_large = m_json['large']
            m_xlarge = m_json['xlarge']
            m_medium = m_json['medium']
            m_close = m_json['close']  # 收盘价
            m_percent = m_json['percent']  # 涨跌幅
            m_amount = m_json['amount'] #主力
            m_time = m_json['timestamp']
            timeStamp = float(m_time / 1000)  # 13位时间戳
            timeArray = time.localtime(timeStamp)
            otherStyleTime = time.strftime("%Y-%m-%d", timeArray)  # 年月日

            if flag == True:
                if new_sheet == False:
                    cell = self.sheet.cell(t_row - 1, 1).value
                    if cell == otherStyleTime:
                        break
                    elif n < 2:
                        self.sheet.cell(row=t_row, column= t_col, value=otherStyleTime)
                        self.sheet.cell(row=t_row, column= t_col + 1, value=m_close)
                        self.sheet.cell(row=t_row, column= t_col + 2, value=str(m_percent) + "%")
                        self.sheet.cell(row=t_row, column= t_col + 3, value=round(m_amount / 10000, 2))
                        self.sheet.cell(row=t_row, column= t_col + 4, value=round(m_xlarge / 10000, 2))
                        self.sheet.cell(row=t_row, column= t_col + 5, value=round(m_large / 10000, 2))
                        self.sheet.cell(row=t_row, column= t_col + 6, value=round(m_medium / 10000, 2))
                        self.sheet.cell(row=t_row, column= t_col + 7, value=round(m_small / 10000, 2))
                        t_row = t_row - 1
                        n = n + 1
                    else:
                        break
                elif new_sheet == True: #新文件/新表 写入全部日期
                    if max_row > 1:
                        self.sheet.cell(row=max_row, column=t_col, value=otherStyleTime)
                        self.sheet.cell(row=max_row, column=t_col + 1, value=m_close)
                        self.sheet.cell(row=max_row, column=t_col + 2, value=str(m_percent) + "%")
                        self.sheet.cell(row=max_row, column= t_col + 3, value=round(m_amount / 10000, 2))
                        self.sheet.cell(row=max_row, column= t_col + 4, value=round(m_xlarge / 10000, 2))
                        self.sheet.cell(row=max_row, column= t_col + 5, value=round(m_large / 10000, 2))
                        self.sheet.cell(row=max_row, column= t_col + 6, value=round(m_medium / 10000, 2))
                        self.sheet.cell(row=max_row, column= t_col + 7, value=round(m_small / 10000, 2))
                        max_row = max_row - 1
            elif flag == False:  # 新文件/新表 写入全部日期
                if max_row > 1:
                    self.sheet.cell(row=max_row, column=t_col, value=otherStyleTime)
                    self.sheet.cell(row=max_row, column=t_col + 1, value=m_close)
                    self.sheet.cell(row=max_row, column=t_col + 2, value=str(m_percent) + "%")
                    self.sheet.cell(row=max_row, column=t_col + 3, value=round(m_amount / 10000, 2))
                    self.sheet.cell(row=max_row, column=t_col + 4, value=round(m_xlarge / 10000, 2))
                    self.sheet.cell(row=max_row, column=t_col + 5, value=round(m_large / 10000, 2))
                    self.sheet.cell(row=max_row, column=t_col + 6, value=round(m_medium / 10000, 2))
                    self.sheet.cell(row=max_row, column=t_col + 7, value=round(m_small / 10000, 2))
                    max_row = max_row - 1

        self.threadLock.release()
        if if_os == True:
            if flag == True:
                try:
                    desktop_path = os.path.join(os.path.expanduser('~'), "Desktop")  # 获取桌面路径
                    path = desktop_path + "\\Finance\\History\\闻讯_主力资金历史.xlsx"
                    self.wb.save(path)
                except:
                    print("Self_Stock Save Error = query_2_1")
            elif flag == False:
                try:
                    desktop_path = os.path.join(os.path.expanduser('~'), "Desktop")
                    path = desktop_path + "\\Finance\\Main_History.xlsx"
                    self.wb.save(path)
                except:
                    print("Self_Stock Save Error = query_2_2")
        elif if_os == False:
            if flag == True:
                try:
                    d_path = "./Finance/History/"
                    path = d_path + "闻讯_主力资金历史.xlsx"
                    self.wb.save(path)
                except:
                    print("Self_Stock Save Error = query_2_1")
            elif flag == False:
                try:
                    file_name = "./Main_History.xlsx"
                    self.wb.save(file_name)
                except:
                    print("Self_Stock Save Error = query_2_2")
    """

    def Deal_Xq_blocktrans(self, data, name):
        self.threadLock.acquire()
        file_name = self.get_filename(name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name(name)

        t_row = sheet.max_row + 3
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="大宗交易")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="成交价")
        sheet.cell(row=t_row, column=t_col + 1, value="成交量(/股)")
        sheet.cell(row=t_row, column=t_col + 2, value="成交额(/万)")
        sheet.cell(row=t_row, column=t_col + 3, value="溢价率")
        sheet.cell(row=t_row, column=t_col + 4, value="交易时间")
        sheet.cell(row=t_row, column=t_col + 5, value="买方营业部")
        sheet.cell(row=t_row, column=t_col + 6, value="卖方营业部")
        t_row = t_row + 1

        m_data = data['data']
        data_items = m_data['items']

        for m_json in data_items:
            m_vol = m_json['vol']
            m_seller = m_json['sell_branch_org_name']
            m_premium = m_json['premium_rat']
            m_trans = m_json['trans_amt']
            m_time = m_json['td_date']
            m_buyer = m_json['buy_branch_org_name']
            m_price = m_json['trans_price']

            timeStamp = float(m_time / 1000)  # 13位时间戳
            timeArray = time.localtime(timeStamp)
            m_date = time.strftime("%Y-%m-%d", timeArray)

            sheet.cell(row=t_row, column=t_col, value=m_price)
            sheet.cell(row=t_row, column=t_col + 1, value=m_vol)
            sheet.cell(row=t_row, column=t_col + 2, value=round(m_trans / 10000, 2))
            sheet.cell(row=t_row, column=t_col + 3, value=str(m_premium) + "%")
            sheet.cell(row=t_row, column=t_col + 4, value=m_date)
            sheet.cell(row=t_row, column=t_col + 5, value=m_seller)
            sheet.cell(row=t_row, column=t_col + 6, value=m_buyer)
            t_row = t_row + 1

        self.threadLock.release()

        try:
            wb.save(file_name)
        except:
            print("Self_Stock Save Error = blocktrans")


    def mkdir_stock(self):
        platform = pt.get_platform()
        if platform == True:
            desktop_path = os.path.join(os.path.expanduser('~'), "Desktop")  # 获取桌面路径
            path = desktop_path + "\\Finance\\Stock\\"
            isExists = os.path.exists(path)
            if not isExists:
                os.mkdir(path)
            return path
        else:
            path = "./Finance/Stock/"
            isExists = os.path.exists(path)
            if not isExists:
                os.mkdir(path)
            return path


    def get_path(self, name):
        platform = pt.get_platform()
        if platform == True:
            desktop_path = os.path.join(os.path.expanduser('~'), "Desktop")  # 获取桌面路径
            path = desktop_path + "\\Finance\\Stock\\{}\\".format(name)
            isExists = os.path.exists(path)
            if not isExists:
                os.mkdir(path)
            return path
        else:
            path = "./Finance/Stock/{}/".format(name)
            isExists = os.path.exists(path)
            if not isExists:
                os.mkdir(path)
            return path


    def get_filename(self, name):
        platform = pt.get_platform()
        if platform == True:
            desktop_path = os.path.join(os.path.expanduser('~'), "Desktop")  # 获取桌面路径
            win_file = desktop_path + "\\Finance\\Stock\\{}\\".format(name) + name + ".xlsx"
            return win_file
        else:
            lin_file = "./Finance/Stock/{}/".format(name) + name + ".xlsx"
            return lin_file


    def Download_Xlsx(self, m_url, path, name, m_time): #这个本来想写入伊利xlsx的 可是格式问题 openpyxl不能load 暂时这样吧
        #filetime = time.strftime("%Y-", time.localtime())  # year-month-day-hour-minute
        filename = path + name + "成交明细_" + m_time + ".xlsx"
        if os.path.exists(filename):
            os.remove(filename)
        for _ in range(3):
            try:
                wget.download(m_url, out=filename)
                break
            except:
                continue

        #filename = path + name + "成交明细_" + "2020_12_18" + ".xlsx" #补充操作

    con = 0
    def Deal_Xq(self, data, name): #忘记这么写的原因了，能跑咱就不动了
        con = self.con
        if self.stop == False:
            return
        if con < 4:
            if con == 0:
                con = con + 1
                t1 = threading.Thread(target=self.Deal_Xq_quote, args=(data, name, ))
                t1.start()
                t1.join()
            elif con == 1:
                con = con + 1
                t2 = threading.Thread(target=self.Deal_Xq_distribution, args=(data, name, ))
                t2.start()
                t2.join()
            elif con == 2:
                con = con + 1
                t3 = threading.Thread(target=self.Deal_Xq_query, args=(data, name, ))
                t3.start()
                t3.join()
            elif con == 3:
                con = con + 1
                t4 = threading.Thread(target=self.Deal_Xq_blocktrans, args=(data, name,))
                t4.start()
                t4.join()
        #elif con < 5:
        self.con = self.con + 1

    def Analysis_date(self, s_url):
        for _ in range(3):
            try:
                self.xlsxdata = requests.get(s_url, headers=headers, timeout=(10, 30))
                break
            except Exception as e:
                continue

        soup = BeautifulSoup(self.xlsxdata.text, "lxml")
        xlslist = soup.select('#historyData > div.bd > a:nth-child(1)')
        for xlsx in xlslist:
            return xlsx.get_text()

    def get_SelfStock(self):
        self.mkdir_stock()
       #url_list = list()
        name_list = dict()
        t = time.time()
        m_time = int(t)
        m_file_time = ''
        with open("Code.txt", "r", encoding='utf-8') as f:
            for line in f.readlines():
                m_line = line.strip('\n')  # 去掉列表中每一个元素的换行符
                if m_line == '':
                    continue
                sep = '#'
                code = line.split(sep, 1)[0]
                if code != '':
                    name = m_line.split(sep)[1]
                    m_code = re.sub('[a-zA-Z]', "", code) #纯数字代码 SH000001 = 000001
                    m_low_code = code.lower() #SH000001 = sh000001
                    m_char = re.sub('[0-9]', "", code) #字母字符  SH0000001 = SH
                    m_low_char = m_char.lower() #SH小写字符
                    #雪球
                    xq_url_quote = 'https://stock.xueqiu.com/v5/stock/quote.json?symbol={0}&extend=detail'.format(code) #个股信息
                    # xq_url_quote = 'https://stock.xueqiu.com/v5/stock/batch/quote.json?extend=detail&is_delay_ft=1&is_delay_hk=0&symbol={}'.format(code) #个股信息
                    xq_url_distrbution = 'https://stock.xueqiu.com/v5/stock/capital/distribution.json?symbol={}&_={}'.format(code, m_time) #今日流出
                    xq_url_query = 'https://stock.xueqiu.com/v5/stock/capital/query.json?count=20&symbol={}&_={}'.format(code, m_time) #流出历史
                    xq_url_blocktrans = 'https://stock.xueqiu.com/v5/stock/capital/blocktrans.json?symbol={}'.format(code) #大宗交易

                    year_filetime = time.strftime("%Y", time.localtime())  # year-month-day-hour-minute
                    #filetime = time.strftime("%Y%m%d", time.localtime())  # year-month-day-hour-minute
                    num = 0
                    if int(m_code) < 600000:
                        num = 1
                    #wyurl = 'http://quotes.money.163.com/cjmx/2021/20210929/1002241.xls'
                    s_url = 'http://quotes.money.163.com/trade/cjmx_{0}.html'.format(m_code)
                    self.filetime = self.Analysis_date(s_url)
                    wyurl = ''
                    if self.filetime != '':
                        m_file_time = self.filetime
                        wyfiletime = self.filetime.replace('-','')
                        #'http://quotes.money.163.com/cjmx/2021/20210809/1300033.xls'
                        wyurl = 'http://quotes.money.163.com/cjmx/{0}/{1}/{2}{3}.xls'.format(year_filetime, wyfiletime, num, m_code)

                    #腾讯证券
                    #tx_url_detail = 'http://stock.gtimg.cn/data/index.php?appn=detail&action=downlddoad&c={}&d={}'.format(m_low_code, filetime) #失效
                    #tx_url_detail = 'http://stock.gtimg.cn/data/index.php?appn=detail&action=download&c={}&d={}'.format(m_low_code, 20201218) #补充日期 失效

                    name_list.setdefault(name, [])
                    name_list[name].append(xq_url_quote)
                    name_list[name].append(xq_url_distrbution)
                    if wyurl != '':
                        name_list[name].append(wyurl)
                    name_list[name].append(xq_url_query)
                    name_list[name].append(xq_url_blocktrans)
                    self.filetime = ''

        url = 'https://xueqiu.com'
        session = requests.session()

        session.get(url, headers=headers)
        for name in name_list:
            self.stop = True
            path = self.get_path(name)
            for m_url in name_list[name]:
                resp = None
                res = "xueqiu" in m_url
                if res == True:
                    for ll in range(3):
                        try:
                            resp = requests.get(m_url, headers=headers, timeout=120)
                            if resp.status_code == 200:
                                break
                        except Exception as e:
                            pass

                    data = json.loads(resp.text)
                    print(data)
                    if data == None:
                        continue

                    self.Deal_Xq(data, name)
                    res = False

                res = "quotes" in m_url
                if res == True:
                    self.Download_Xlsx(m_url, path, name, m_file_time)
                    res = False
                time.sleep(1)  # 减速
            self.con = 0

        del name_list #释放
        self.filetime = ''


    def main(self, filename):
        #Stock = SelfStock(file_name)
        #Stock.get_filename()
        Stock.get_SelfStock()

Stock = SelfStock()
