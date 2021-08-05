import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from queue import Queue

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class SongGuo(object):
    def __init__(self):
        pass

    def request(self):
        self.url = 'http://www.songguocaijing.com/'
        self.data = requests.get(self.url, headers=headers)
        self.data.encoding = "utf-8"
        self.soup = BeautifulSoup(self.data.text, "lxml")


    def get_news(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet("Sg")
        t_row = 1
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="松果财经")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist = self.soup.find_all(class_='title-wp')
        da = ''
        da2 = ''
        da2 = self.soup.find_all(class_='article-description')
        m_queue = Queue(20)
        for des in da2:
            m_queue.put(des.get_text())

        for da in datalist:
            da.find('a')
            for news in da:
                m_href = 'http://www.songguocaijing.com/' + news['href']
                m_title = news['title']
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                sheet.cell(row=t_row, column=t_col + 2, value=m_queue.get())
                t_row = t_row + 1

        try:
            wb.save(self.xlsxname)
        except:
            print("Songguo Save Error")

    def main(self, file_name):
        self.xlsxname = file_name
        Sg.request()
        Sg.get_news()

Sg = SongGuo()
