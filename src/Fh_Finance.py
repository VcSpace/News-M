import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import re
import json
import time

"""
https://tech.ifeng.com/24h/
http://tech.ifeng.com/
http://finance.ifeng.com/
"""

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class FengHuang(object):
    def __init__(self):
        pass

    def request(self):
        #new
        self.url = 'http://finance.ifeng.com/'
        self.data = requests.get(self.url, headers=headers)
        self.soup = BeautifulSoup(self.data.text, "lxml")

        self.stock_url = 'http://finance.ifeng.com/stock/'
        self.stock_data = requests.get(self.stock_url, headers=headers)
        self.stock_soup = BeautifulSoup(self.stock_data.text, "lxml")


    def getTopNew(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet("Fh")
        t_row = 1
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="凤凰财经")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        #datalist = self.soup.find_all(class_='hot-1NJ2DKa4 clearfix')
        datalist = self.soup.select('#root > div > div.col-3u4gcc0Q.clearfix > div.col_L-3c5atSII > div.box-1bAs3EGr')

        for Newslist in datalist:
            News = Newslist.find_all('a')
            for m_new in News:
                m_title = m_new['title']
                m_href = m_new['href']
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("FengHuang Save Error = getTopNew")


    def stockNewslist(self, json_str, t_row): #stockNewsList
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Fh")
        Newslist = json_str['stockNewsList']
        t_col = 1
        temp = 0
        for m_new in Newslist:
            temp = temp + 1
            if temp == 6:
                temp = 0
                t_row = t_row + 1
            else:
                m_title = m_new['title']
                m_time = m_new['newsTime']
                m_href = m_new['url']
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                sheet.cell(row=t_row, column=t_col + 3, value=m_time)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("FengHuang Save Error = getNewsList")


    def CompanyNews(self, json_str): #stockNewsList 没有top  只有top下的5条
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Fh")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="公司要闻")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        CNews = json_str['news']
        m_CNews = json_str['newsList']
        for m_new in CNews:
            m_href = m_new['url']
            m_title = m_new['title']
            m_href = m_href.replace("//", "https://")
            sheet.cell(row=t_row, column=t_col, value=m_title)
            sheet.cell(row=t_row, column=t_col + 1, value=m_href)
            t_row = t_row + 1
            break

        for m_new in m_CNews:
            m_time = m_new['newsTime']
            m_href = m_new['url']
            m_title = m_new['title']
            sheet.cell(row=t_row, column=t_col, value=m_title)
            sheet.cell(row=t_row, column=t_col + 1, value=m_href)
            sheet.cell(row=t_row, column=t_col + 2, value=m_time)
            t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("FengHuang Save Error = CompanyNews2")

    def marketAnalysis(self, json_str):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Fh")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="操盘分析")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        Newslist = json_str['marketAnalysis']
        for m_news in Newslist:
            m_href = m_news['url']
            m_title = m_news['title']
            m_href = m_href.replace("//", "https://")
            sheet.cell(row=t_row, column=t_col, value=m_title)
            sheet.cell(row=t_row, column=t_col + 1, value=m_href)
            t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("FengHuang Save Error = marketAnalysis")


    def IPOObservation(self, json_str): #stockNewsList 没有top  只有top下的5条
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Fh")
        t_row = sheet.max_row + 2
        t_col = 1

        sheet.cell(row=t_row, column=t_col, value="IPO观察")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        NewsList = json_str['hotPlate']
        for m_new in NewsList:
            m_href = m_new['url']
            m_title = m_new['title']
            m_href = m_href.replace("//", "https://")
            sheet.cell(row=t_row, column=t_col, value=m_title)
            sheet.cell(row=t_row, column=t_col + 1, value=m_href)
            t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("FengHuang Save Error = IPOObservation")


    def getStockNews(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Fh")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="证券要闻")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1
        soup = str(self.stock_soup)

        #sear = re.search(r'var allData = (.*?) .*\"\}\}\;', soup, re.M | re.I)
        sear = re.search('var allData = (.*?)var adData', soup, re.M | re.I | re.S)
        data = sear.group(1)
        data = data.replace(";", "")
        json_str = json.loads(data)
        #print(data)
        m_row = t_row + 1
        stockNews = json_str['stockNews'] #top新闻
        for m_new in stockNews:
            m_title = m_new['title']
            href = m_new['url']
            m_href = href.replace("//", "https://")
            sheet.cell(row=t_row, column=t_col, value=m_title)
            sheet.cell(row=t_row, column=t_col + 1, value=m_href)
            t_row = t_row + 1

        try:
            wb.save(self.xlsxname)
        except:
            print("FengHuang Save Error = getStockNews")

        self.stockNewslist(json_str, m_row) #stockNewsList 没有top  只有top下的5条
        self.CompanyNews(json_str) #stockNewsList 没有top  只有top下的5条
        self.marketAnalysis(json_str)
        self.IPOObservation(json_str) #stockNewsList 没有top  只有top下的5条

    def main(self, filename):
        self.xlsxname = filename
        #Fh.Style()
        Fh.request()
        Fh.getTopNew()
        Fh.getStockNews()

Fh = FengHuang()
