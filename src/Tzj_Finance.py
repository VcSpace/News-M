import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15',
}

class Touzijie(object):
    def __init__(self):
        pass

    def request(self):
        self.url = 'https://www.pedaily.cn'
        self.data = requests.get(self.url, headers=headers,timeout=60)
        self.data.encoding = "utf-8"
        self.soup = BeautifulSoup(self.data.text, "lxml")

    def get_topnews(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet("Tzj")
        t_row = 1
        t_col = 1

        sheet.cell(row=t_row, column=t_col, value="每日TOP5")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist = self.soup.select('#box-fix-content > div.tab-content > ul:nth-child(1)')
        for data in datalist:
            news = data.find_all('a')
            for m_new in news:
                m_href = m_new['href']
                m_title = m_new.get_text()
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1

        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Tzj Save Error = 1")

    def get_news(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Tzj")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="最新资讯")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1
        datalist = self.soup.find_all(class_='news-list news-list-bottom news-list-special')
        for data in datalist:
            news = data.find_all(class_='txt')
            for m_news in news:
                m_news = m_news.find_all('h3')
                pattern = re.compile(r'((https|http)?:\/\/)[^\s]+[^" t]')
                searchObj = re.search(r'(.*)target="_blank">(.*?)</a>', str(m_news), re.M | re.I)
                m_href = pattern.search(str(m_news)).group()
                m_title = searchObj.group(2)
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Tzj Save Error = 2")

    def get_Instantnews(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Tzj")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="即时快讯")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1
        datalist = self.soup.find_all(class_='list-time hot-online')

        for trtag in datalist:
            data = trtag.find_all('li')  # 在每个tr标签下,查找所有的td标签
            for news in data:
                m_href = news['data-url']
                m_title = news['data-title']
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1

        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Tzj Save Error = 2")

    def get_invest(self):
        datalist = self.soup.find_all(class_='list-invest')
        print(datalist)

    def get_ipo(self):
        datalist = self.soup.find_all(class_='list-ipo')
        print(datalist)


    def main(self, file_name):
        self.xlsxname = file_name
        try:
            Tzj.request()
            Tzj.get_topnews()
            Tzj.get_news()
            Tzj.get_Instantnews()
        #Tzj.get_invest()
        #Tzj.get_ipo()
        except Exception:
            pass


Tzj = Touzijie()
