import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class XinHuaNet(object):
    def __init__(self):
        pass

    def request(self):
        self.url = 'http://www.xinhuanet.com/fortunepro/'
        for ll in range(3):
            try:
                self.data = requests.get(self.url, headers=headers, timeout=120)
                if self.data.status_code == 200:
                    break
            except Exception as e:
                pass

        self.data.encoding = "utf-8"
        self.soup = BeautifulSoup(self.data.text, "lxml")


    def getTopNew(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet("Xhs")
        t_row = 1
        t_col = 1

        sheet.cell(row=t_row, column=t_col, value="财经TOP10")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist = self.soup.find_all(class_='cjtop')
        for newlist in datalist:
            news = newlist.find_all('a')
            for m_new in news:
                m_href = m_new['href']
                m_title = m_new.get_text()
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1

        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Xhs Save Error = 1")

    def getnews(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Xhs")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="财经新闻")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1
        datalist = self.soup.find_all(class_='xpage-content-list')
        for data in datalist:
            news = data.find_all('a')
            for m_new in news:
                m_href = m_new['href']
                m_title = m_new.get_text()
                if m_title == "":
                    continue
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Xhs Save Error = 2")


    def get_research_report(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Xhs")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="行业研报")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1
        datalist = self.soup.find_all(class_='rtlist')
        for data in datalist:
            news = data.find_all('a')
            for m_new in news:
                m_href = m_new['href']
                m_title = m_new.get_text()
                if len(m_title) <= 4:
                    continue
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1

        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Xhs Save Error = 3")


    def main(self, file_name):
        self.xlsxname = file_name
        Xhs.request()
        Xhs.getTopNew()
        Xhs.getnews()
        Xhs.get_research_report()


Xhs = XinHuaNet()
