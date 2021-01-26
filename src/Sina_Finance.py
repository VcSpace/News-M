import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime
import time
import json
import threading

"""
https://finance.sina.com.cn/
http://finance.sina.com.cn/chanjing/
http://feed.mix.sina.com.cn/api/roll/get?pageid=164&lid=1694&num=10&page=1&callback=feedCardJsonpCallback&_=1611561556802 公司新闻
http://feed.mix.sina.com.cn/api/roll/get?pageid=164&lid=1695&num=10&page=1&callback=feedCardJsonpCallback&_=1611561513495 产业新闻

http://finance.sina.com.cn/china/
http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1687&num=10&page=1&callback=feedCardJsonpCallback&_=1611573471132 国内新闻
http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1687&num=10&page=1&callback=feedCardJsonpCallback&_=1611573513741 宏观经济
http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1689&num=10&page=1&callback=feedCardJsonpCallback&_=1611573537003 部委动态
http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1690&num=10&page=1&callback=feedCardJsonpCallback&_=1611573555700 金融新闻
http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1688&num=10&page=1&callback=feedCardJsonpCallback&_=1611573598249 地方经济
"""

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class SinaNews(object):
    def __init__(self):
        pass

    def request(self):
        self.url = 'https://finance.sina.com.cn/'
        self.data = requests.get(self.url, headers=headers)
        self.data.encoding = "utf-8"
        self.soup = BeautifulSoup(self.data.text, "lxml")


    def getTopNew(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet("Sina")
        t_row = 1
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="新浪财经")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist = self.soup.find_all(class_='fin_tabs0_c0')
        for datal in datalist:
            t = 0 #留着先

        for ds in datal:
            ds = datal.find_all('a')
            for m_data in ds:
                m_href = m_data['href']
                m_title = m_data.get_text()
                if len(m_title) < 4:
                    continue
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1
            break

        try:
            wb.save(self.xlsxname)
        except:
            print("SinaNews Save Error = getTopNew")

    def getStockNew(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Sina")
        t_row = sheet.max_row + 3
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="证券新闻")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist = self.soup.find_all(class_='m-p1-m-blk2')
        for dastl in datalist:
            t = 1

        for ds in dastl:
            ds = dastl.find_all('a')
            for m_data in ds:
                m_href = m_data['href']
                m_title = m_data.get_text()
                if len(m_title) < 4:
                    continue
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1
            break

        try:
            wb.save(self.xlsxname)
        except:
            print("SinaNews Save Error = getStockNew")


    def getIndustryNew(self):

        t = time.time() * 1000
        n_time = int(t)
        new_list = list()

        j_url1 = 'http://feed.mix.sina.com.cn/api/roll/get?pageid=164&lid=1694&num=10&page=1&callback=feedCardJsonpCallback&_={}'.format(n_time) #公司新闻
        j_url2 = 'http://feed.mix.sina.com.cn/api/roll/get?pageid=164&lid=1695&num=10&page=1&callback=feedCardJsonpCallback&_={}'.format(n_time) #产业新闻

        cn_url1 = 'http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=3231&num=10&page=1&callback=feedCardJsonpCallback&_={}'.format(n_time) #财经top10
        cn_url2 = 'http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1686&num=10&page=1&callback=feedCardJsonpCallback&_={}'.format(n_time) #国内新闻
        cn_url3 = 'http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1687&num=10&page=1&callback=feedCardJsonpCallback&_={}'.format(n_time) #宏观经济
        cn_url4 = 'http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1689&num=10&page=1&callback=feedCardJsonpCallback&_={}'.format(n_time) #部委动态
        cn_url5 = 'http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1690&num=10&page=1&callback=feedCardJsonpCallback&_={}'.format(n_time) #金融新闻
        cn_url6 = 'http://feed.mix.sina.com.cn/api/roll/get?pageid=155&lid=1688&num=10&page=1&callback=feedCardJsonpCallback&_={}'.format(n_time) #地方新闻
        new_list = {
            "财经Top10": cn_url1,
            "公司新闻": j_url1,
            "产业新闻": j_url2,
            "国内新闻": cn_url2,
            "宏观经济": cn_url3,
            "部委动态": cn_url4,
            "金融新闻": cn_url5,
            "地方新闻": cn_url6,
        }
        for newname in new_list:
            m_url = new_list[newname]
            self.Deal_News(m_url, newname)


    def Deal_News(self, url, new_name):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name("Sina")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="{}".format(new_name))
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        data = requests.get(url, headers=headers)
        t1 = '<html><body><p>try{feedCardJsonpCallback('
        t2 = ');}catch(e){};</p></body></html>'
        soup = BeautifulSoup(data.text, "lxml")
        data = str(soup)
        data = data.replace(t1, "")
        data = data.replace(t2, "")
        json_str = json.loads(data)
        json_str = json_str['result']['data']
        for m_data in json_str:
            m_href = m_data['url']
            m_title = m_data['title']
            tmp_m_time = m_data['ctime']
            timeStamp = int(tmp_m_time)
            timeArray = time.localtime(timeStamp)
            m_time = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)

            sheet.cell(row=t_row, column=t_col, value=m_title)
            sheet.cell(row=t_row, column=t_col + 1, value=m_href)
            sheet.cell(row=t_row, column=t_col + 3, value=m_time)
            t_row = t_row + 1


        """
        t_row = t_row + 2
        sheet.cell(row=t_row, column=t_col, value="产业新闻")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        data = requests.get(j_url2, headers=headers)
        soup = BeautifulSoup(data.text, "lxml")
        data = str(soup)
        m1 = '<html><body><p>try{feedCardJsonpCallback('
        m2 = ');}catch(e){};</p></body></html>'
        data = data.replace(m1, "")
        data = data.replace(m2, "")
        json_str2 = json.loads(data)
        json_s = json_str2['result']['data']
        for m_data in json_s:
            m_href = m_data['url']
            m_title = m_data['title']
            tmp_m_time = m_data['ctime']
            timeStamp = int(tmp_m_time)
            timeArray = time.localtime(timeStamp)
            m_time = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)

            sheet.cell(row=t_row, column=t_col, value=m_title)
            sheet.cell(row=t_row, column=t_col + 1, value=m_href)
            sheet.cell(row=t_row, column=t_col + 3, value=m_time)
            t_row = t_row + 1
        """
        try:
            wb.save(self.xlsxname)
        except:
            print("Sina getIndustryNew Save Error")

    def getCnNew(self):

        t = time.time() * 1000
        n_time = int(t)



    def main(self, file_name):
        self.xlsxname = file_name
        Sina.request()
        Sina.getTopNew()
        Sina.getStockNew()
        Sina.getIndustryNew()
        Sina.getCnNew()

Sina = SinaNews()
