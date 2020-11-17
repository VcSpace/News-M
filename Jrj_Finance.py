import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
import json
import time
import datetime

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class JinRongJie(object):
    def __init__(self, filename):
        self.xlsxname = filename

    def Style(self):
        self.m_font = Font(
            size=12,
            bold=True,
        )

        self.head_font = Font(
            size=14,
            bold=True,
        )

    def get_TopNews(self):
        url = 'http://finance.jrj.com.cn/'
        data = requests.get(url, headers=headers)
        soup = BeautifulSoup(data.text, "lxml")

        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet('Jrj')
        t_row = 1
        t_col = 1

        sheet.cell(row=t_row, column=t_col + 0, value="金融界财经")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col + 0, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist = soup.find_all(class_="l1_top")
        for News in datalist:
            New = News.select('dl dt p')
            for m_new in New:
                data = m_new.find('a')
                m_url = data['href']
                m_title = data.get_text()
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("JRJ Save Error = 1")

    def get_FinanceNews(self): #差不多4-5个小时内的热点新闻
        fin_time1 = time.strftime("%Y%m", time.localtime())  # year-month-day-hour-minute
        fin_time2 = time.strftime("%Y%m%d", time.localtime())  # year-month-day-hour-minute
        fin_url = 'http://finance.jrj.com.cn/xwk/{}/{}_1.shtml'.format(fin_time1, fin_time2)

        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Jrj')
        t_row = sheet.max_row + 1
        t_col = 1

        sheet.cell(row=t_row + 1, column=t_col, value="财经频道新闻")
        sheet.cell(row=t_row + 2, column=t_col, value="新闻标题")
        sheet.cell(row=t_row + 2, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row + 2, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row + 2, column=t_col + 3, value="新闻时间")
        t_row = t_row + 3
        time_row = t_row

        data = requests.get(fin_url, headers=headers)
        soup = BeautifulSoup(data.text, "lxml")
        datalist = soup.find_all(class_="list")
        flag = False
        for Newslist in datalist:
            #News = Newslist.select('li')
            News = Newslist.find_all('a')
            m_NewsTime = Newslist.find_all('span')
            for m_new in News:
                if flag == True:
                    flag = False
                    m_url = m_new['href']
                    m_title = m_new.get_text()
                    sheet.cell(row=t_row, column=t_col, value=m_title)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                    t_row = t_row + 1
                else:
                    flag = True
            for new_time in m_NewsTime:
                m_time = new_time.get_text()
                sheet.cell(row=time_row, column=t_col + 3, value=m_time)
                time_row = time_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("JRJ Save Error = 2")

    def get_todayHot(self):
        url = 'http://biz.jrj.com.cn/biz_index.shtml'
        data = requests.get(url, headers=headers)
        soup = BeautifulSoup(data.text, "lxml")
        datalist = soup.find_all(class_="jrj-top10")

        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Jrj')
        t_row = sheet.max_row + 2
        t_col = 1

        sheet.cell(row=t_row, column=t_col, value="24小时间热闻点击排行榜")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 0, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        flag = True
        for Newslist in datalist:
            if flag == True:
                News = Newslist.find_all('a')
                for m_new in News:
                    m_title = m_new['title']
                    m_url = m_new['href']
                    sheet.cell(row=t_row, column=t_col, value=m_title)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                    t_row = t_row + 1
            flag = False
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("JRJ Save Error = 5")


    def get_Business(self):
        bu_time1 = time.strftime("%Y%m", time.localtime())  # year-month-day-hour-minute
        bu_time2 = time.strftime("%Y%m%d", time.localtime())  # year-month-day-hour-minute
        bus_url = 'http://biz.jrj.com.cn/xwk/{}/{}_1.shtml'.format(bu_time1, bu_time2)

        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Jrj')
        t_row = sheet.max_row + 1
        t_col = 1

        sheet.cell(row=t_row + 1, column=t_col, value="商业频道新闻")
        sheet.cell(row=t_row + 2, column=t_col, value="新闻标题")
        sheet.cell(row=t_row + 2, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row + 2, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row + 2, column=t_col + 3, value="新闻时间")
        t_row = t_row + 3
        time_row = t_row

        data = requests.get(bus_url, headers=headers)
        soup = BeautifulSoup(data.text, "lxml")
        datalist = soup.find_all(class_="list")
        flag = False
        for Newslist in datalist:
            # News = Newslist.select('li')
            News = Newslist.find_all('a')
            m_NewsTime = Newslist.find_all('span')
            for m_new in News:
                if flag == True:
                    flag = False
                    m_url = m_new['href']
                    m_title = m_new.get_text()
                    sheet.cell(row=t_row, column=t_col, value=m_title)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                    t_row = t_row + 1
                else:
                    flag = True
            for new_time in m_NewsTime:
                m_time = new_time.get_text()
                sheet.cell(row=time_row, column=t_col + 3, value=m_time)
                time_row = time_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("JRJ Save Error = 3")

    """
    def get_Science(self):
        url = 'http://finance.jrj.com.cn/tech/tech_index.shtml'
        sci_time = time.strftime("%Y-%m-%d", time.localtime())  # year-month-day-hour-minute
    """

    def get_yesScience(self):
        sci_time1 = time.strftime("%Y%m", time.localtime())  # year-month-day-hour-minute
        sci_time2 = time.strftime("%Y%m%d", time.localtime())  # year-month-day-hour-minute
        """ 
        yesterday = time.strftime("%d", time.localtime())  # year-month-day-hour-minute
        yesterday = 1
        if yesterday == 1:
            # 获取当前日期
            now_time = datetime.datetime.now()
            # 获取本月的第一天
            end_day_in_mouth = now_time.replace(day=1)
            # 获取上月的最后一天
            next_mouth = end_day_in_mouth - datetime.timedelta(days=1)
            print(next_mouth.month)
        """
        sci_url = 'http://biz.jrj.com.cn/xwk/{}/{}_1.shtml'.format(sci_time1, sci_time2)

        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Jrj')
        t_row = sheet.max_row
        t_col = 1

        sheet.cell(row=t_row + 1, column=t_col, value="科技频道新闻")
        sheet.cell(row=t_row + 2, column=t_col, value="新闻标题")
        sheet.cell(row=t_row + 2, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row + 2, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row + 2, column=t_col + 3, value="新闻时间")
        t_row = t_row + 3
        time_row = t_row

        data = requests.get(sci_url, headers=headers)
        soup = BeautifulSoup(data.text, "lxml")
        datalist = soup.find_all(class_="list")
        flag = False
        for Newslist in datalist:
            # News = Newslist.select('li')
            News = Newslist.find_all('a')
            m_NewsTime = Newslist.find_all('span')
            for m_new in News:
                if flag == True:
                    flag = False
                    m_url = m_new['href']
                    m_title = m_new.get_text()
                    sheet.cell(row=t_row, column=t_col, value=m_title)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                    t_row = t_row + 1
                else:
                    flag = True
            for new_time in m_NewsTime:
                m_time = new_time.get_text()
                sheet.cell(row=time_row, column=t_col + 3, value=m_time)
                time_row = time_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("JRJ Save Error = 4")

    def main(self, filename):
        Jrj = JinRongJie(filename)
        Jrj.Style()
        Jrj.get_TopNews()
        Jrj.get_todayHot()
        Jrj.get_FinanceNews()
        Jrj.get_Business()
#        Jrj.get_Science()
        Jrj.get_yesScience()

