import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
import json
#from requests.cookies import RequestsCookieJar
import threading
import time
import re

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
    #'Cookie': 'xqat=3e14cc861fdd960a5d84e7316165286b1bfeafe3;',
}

class SelfStock(object):
    threadLock = threading.Lock()
    def __init__(self, file_name):
        self.xlsxname = file_name

    def Style(self):
        self.m_font = Font(
            size=12,
            bold=True,
        )
        self.head_font = Font(
            size=14,
            bold=True,
        )

    def Deal_Xq_quote(self, data, url, name):
        self.threadLock.acquire()
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet(name)
        t_row = 1
        t_col = 1

        data_json = data['data']
        data_items = data_json['items']
        data_mk = data_items[0]['market']
        data_quote = data_items[0]['quote']
        #print(data_quote)

        sheet.cell(row=t_row, column=t_col, value="股票代码")
        sheet.cell(row=t_row, column=t_col + 1, value="股票名称")
        sheet.cell(row=t_row, column=t_col + 2, value="交易状态")
        sheet.cell(row=t_row, column=t_col + 3, value="更新时间")
        t_row = t_row + 1
        m_status = data_mk['status']
        stock_code = data_quote['symbol']
        stock_name = data_quote['name']
        m_time = data_quote['timestamp']
        timeStamp = float(m_time / 1000) #13位时间戳
        timeArray = time.localtime(timeStamp)
        otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)

        sheet.cell(row=t_row, column=t_col, value=stock_code)
        sheet.cell(row=t_row, column=t_col + 1, value=stock_name)
        sheet.cell(row=t_row, column=t_col + 2, value=m_status)
        sheet.cell(row=t_row, column=t_col + 3, value=otherStyleTime)
        t_row = t_row + 2

        sheet.cell(row=t_row, column=t_col + 0, value="当前价格")
        sheet.cell(row=t_row, column=t_col + 1, value="涨跌幅度")
        sheet.cell(row=t_row, column=t_col + 2, value="涨跌价格")
        sheet.cell(row=t_row, column=t_col + 3, value="开盘价格")
        sheet.cell(row=t_row, column=t_col + 4, value="当前新高")
        sheet.cell(row=t_row, column=t_col + 5, value="当前新低")
        sheet.cell(row=t_row, column=t_col + 6, value="昨日收盘")
        sheet.cell(row=t_row, column=t_col + 7, value="平均价格")
        sheet.cell(row=t_row, column=t_col + 8, value="涨停价格")
        sheet.cell(row=t_row, column=t_col + 9, value="跌停价格")
        sheet.cell(row=t_row, column=t_col + 10, value="52周最高")
        sheet.cell(row=t_row, column=t_col + 11, value="52周最低")
        t_row = t_row + 1

        m_current = data_quote['current']
        m_percent = data_quote['percent'] #涨跌幅度
        m_chg = data_quote['chg'] #涨跌价格
        m_open = data_quote['open'] #开盘价
        m_yesclose = data_quote['last_close'] #昨收
        m_high = data_quote['high']
        m_low = data_quote['low']
        m_avg_price = data_quote['avg_price']
        m_limit_up = data_quote['limit_up'] #涨停
        m_limit_down = data_quote['limit_down'] #跌停
        m_high52w = data_quote['high52w'] #52周最高
        m_low52w = data_quote['low52w'] #52周最低
        sheet.cell(row=t_row, column=t_col + 0, value=m_current)
        sheet.cell(row=t_row, column=t_col + 1, value=m_percent)
        sheet.cell(row=t_row, column=t_col + 2, value=m_chg)
        sheet.cell(row=t_row, column=t_col + 3, value=m_open)
        sheet.cell(row=t_row, column=t_col + 4, value=m_high)
        sheet.cell(row=t_row, column=t_col + 5, value=m_low)
        sheet.cell(row=t_row, column=t_col + 6, value=m_yesclose)
        sheet.cell(row=t_row, column=t_col + 7, value=m_avg_price)
        sheet.cell(row=t_row, column=t_col + 8, value=m_limit_up)
        sheet.cell(row=t_row, column=t_col + 9, value=m_limit_down)
        sheet.cell(row=t_row, column=t_col + 10, value=m_high52w)
        sheet.cell(row=t_row, column=t_col + 11, value=m_low52w)
        t_row = t_row + 2


        sheet.cell(row=t_row, column=t_col + 0, value="成交额")
        sheet.cell(row=t_row, column=t_col + 1, value="成交量")
        sheet.cell(row=t_row, column=t_col + 2, value="换手率")
        sheet.cell(row=t_row, column=t_col + 3, value="量比")
        sheet.cell(row=t_row, column=t_col + 4, value="振幅")
        sheet.cell(row=t_row, column=t_col + 5, value="市盈率TTM")
        sheet.cell(row=t_row, column=t_col + 6, value="市盈率(动)")
        sheet.cell(row=t_row, column=t_col + 7, value="市盈率(静)")
        sheet.cell(row=t_row, column=t_col + 8, value="市净率")
        sheet.cell(row=t_row, column=t_col + 9, value="总股本")
        sheet.cell(row=t_row, column=t_col + 10, value="流通股")
        sheet.cell(row=t_row, column=t_col + 11, value="总市值")
        sheet.cell(row=t_row, column=t_col + 12, value="流通市值")
        t_row = t_row + 1

        m_amount = data_quote['amount'] #成交额
        m_turnover_rate = data_quote['turnover_rate'] #换手：0.74%
        m_volume = data_quote['volume'] #成交量
        m_amplitude = data_quote['amplitude'] #振幅
        m_total_shares = data_quote['total_shares'] #总股本
        m_float_shares = data_quote['float_shares'] #流通股
        m_volume_ratio = data_quote['volume_ratio'] #量比
        m_pe_ttm = data_quote['pe_ttm'] #市盈率
        m_pe_forecast = data_quote['pe_forecast'] #动态市盈率
        m_pe_lyr = data_quote['pe_lyr'] #静态市盈率
        m_pb = data_quote['pb'] #市净率
        m_profit = data_quote['profit'] #年报利润
        m_profit_four = data_quote['profit_four'] #也是利润 不清楚
        m_profit_forecast = data_quote['profit_forecast']
        m_market_capital = data_quote['market_capital'] #总市值
        m_float_market_capital = data_quote['float_market_capital'] #流通市值
        sheet.cell(row=t_row, column=t_col + 0, value=m_amount)
        sheet.cell(row=t_row, column=t_col + 1, value=m_volume)
        sheet.cell(row=t_row, column=t_col + 2, value=str(m_turnover_rate) + "%")
        sheet.cell(row=t_row, column=t_col + 3, value=m_volume_ratio)
        sheet.cell(row=t_row, column=t_col + 4, value=str(m_amplitude) + "%")
        sheet.cell(row=t_row, column=t_col + 5, value=m_pe_ttm)
        sheet.cell(row=t_row, column=t_col + 6, value=m_pe_forecast)
        sheet.cell(row=t_row, column=t_col + 7, value=m_pe_lyr)
        sheet.cell(row=t_row, column=t_col + 8, value=m_pb)
        sheet.cell(row=t_row, column=t_col + 9, value=m_total_shares)
        sheet.cell(row=t_row, column=t_col + 10, value=m_float_shares)
        sheet.cell(row=t_row, column=t_col + 11, value=m_market_capital)
        sheet.cell(row=t_row, column=t_col + 12, value=m_float_market_capital)
        try:
            wb.save(self.xlsxname)
            self.threadLock.release()
        except Exception:
            print("Self_Stock Save Error = Xq_qupte")
            self.threadLock.release()

    #def get_Main_capital_history(self): #主力资金流向历史记录  记录是否跑路
    def Deal_Xq(self, data):
        for i in range(3):



    def get_SelfStock(self):
        #url_list = list()
        name_list = dict()
        t = time.time()
        m_time = int(t)
        with open("Code.txt", "r") as f:
            for line in f.readlines():
                m_line = line.strip('\n')  # 去掉列表中每一个元素的换行符
                sep = '#'
                code = line.split(sep, 1)[0]
                if code != '':
                    name = m_line.split(sep)[1]
                    m_code = re.sub('[a-zA-Z]', "", code)
                    url_code = 'https://stock.xueqiu.com/v5/stock/batch/quote.json?extend=detail&is_delay_ft=1&is_delay_hk=0&symbol={}'.format(code)
                    url_mainc = 'https://stock.xueqiu.com/v5/stock/capital/distribution.json?symbol={}&_={}'.format(code, m_time) #今日流出
                    url_main_h = 'https://stock.xueqiu.com/v5/stock/capital/query.json?count=20&symbol={}&_={}'.format(code, m_time) #流出历史
                    #name_list[str(name)] = url_list
                    name_list.setdefault(name, [])
                    name_list[name].append(url_code)
                    name_list[name].append(url_mainc)
                    name_list[name].append(url_main_h)

        url = 'https://xueqiu.com'
        session = requests.session()
        session.get(url, headers=headers)
        for name in name_list:
            for m_url in name_list[name]:
                res = "xueqiu" in m_url
                if res == True:
                    resp = session.get(m_url, headers=headers)
                    data = json.loads(resp.text)
                    self.Deal_Xq(data)

            """
            resp = session.get(m_url, headers=headers)
            data = json.loads(resp.text)
            t1 = threading.Thread(target=self.Deal_Xq_data, args=(data, url, name, ))
            #t2 = threading.Thread(target=self.get_Main_capital_history, args=(data, url, name, ))
            t1.start()
            t1.join()
            break
            """

        del name_list #释放

    def main(self, file_name):
        Stock = SelfStock(file_name)
        Stock.get_SelfStock()
